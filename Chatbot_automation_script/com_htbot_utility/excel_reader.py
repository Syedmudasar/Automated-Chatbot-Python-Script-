import openpyxl
from com_hrbot_constant.constant import Constant
from openpyxl import load_workbook


# noinspection PyBroadException
class ExcelReader:

    @staticmethod
    def read_excel():

        # workbook object is created
        try:
            wb_obj = openpyxl.load_workbook(Constant.excel_sheet)
        except FileNotFoundError as f:
            print(f)

        sheets = wb_obj.sheetnames
        print(sheets)
        sheet_obj = wb_obj[sheets[0]]
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row

        # Loop will convert excel file data as dictionary.
        # first column will be treated as keys and second one as values

        urls_dict = {}
        for i in range(1, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1, )
            urls_dict[cell_obj.value] = sheet_obj.cell(row=i, column=2).value
        return urls_dict

    # read a particular sheet from excel file

    @staticmethod
    def read_excel1():
        try:
            wb = openpyxl.load_workbook(Constant.excel_sheet)
        except Exception as e:
            print(e)

        sheets = wb.sheetnames
        wb = wb[sheets[2]]

        max_col = wb.max_column
        max_row = wb.max_row

        print(max_col)
        print(max_row)

        urls_dict = {}
        for i in range(1, max_row + 1):
            cell_obj = wb.cell(row=i, column=1)
            urls_dict[cell_obj.value] = wb.cell(row=i, column=2).value
        return urls_dict

    @staticmethod
    def read_excel2():
        try:
            wb = openpyxl.load_workbook(Constant.excel_sheet)
        except Exception as e:
            print(e)

        sheets = wb.sheetnames
        wb = wb[sheets[1]]

        max_col = wb.max_column
        max_row = wb.max_row

        print(max_col)
        print(max_row)

        urls_dict = {}
        for i in range(1, max_row + 1):
            cell_obj = wb.cell(row=i, column=1)
            urls_dict[cell_obj.value] = wb.cell(row=i, column=2).value
        return urls_dict


